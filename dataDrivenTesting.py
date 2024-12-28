from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep


# Load the Excel sheet
workbook = load_workbook("testdata.xlsx")
sheet = workbook.active #Get the current active sheet, ie sheet with data or None
url = "https://www.saucedemo.com/"
browser = webdriver.Chrome()
browser.maximize_window()


# Iterate through the active sheet by indicating the min and maximum rows to iterate
# Also, only rows with value
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    username = row[0]
    password = row[1]

    browser.get(url)
    sleep(5)
    browser.find_element(By .ID, value="user-name").send_keys(username)
    browser.find_element(By .ID, value="password").send_keys(password)

    browser.find_element(By .ID, value="login-button").click()
    sleep(5)

    # Handle logout for each of the login
    browser.find_element(By .XPATH, value="//button[@id='react-burger-menu-btn']").click()
    sleep(3)
    browser.find_element(By .XPATH, value="//a[@id='logout_sidebar_link']").click()
    sleep(3)

# browser.quit()
